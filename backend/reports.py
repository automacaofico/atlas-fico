import hashlib
import io
import json
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from xml.sax.saxutils import escape

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.graphics.shapes import Drawing, Rect, String
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Image, KeepTogether, LongTable, PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

BLUE = "3578BC"
DARK = "18324D"
ORANGE = "F89B32"
LIGHT = "E9F2F9"
ROOT = Path(__file__).resolve().parents[1]
LOGO = ROOT / "assets" / "brand" / "fico-logo-color.png"


def text(value):
    return "" if value is None else str(value)


def safe(value):
    return escape(text(value))


def date_text(value):
    if not value:
        return "-"
    raw = str(value)[:10]
    try:
        return datetime.strptime(raw, "%Y-%m-%d").strftime("%d/%m/%Y")
    except ValueError:
        return str(value)


def history_comment(value):
    if not value:
        return "-"
    raw = str(value)
    try:
        changes = json.loads(raw)
        if isinstance(changes, dict) and changes:
            lines = []
            for field, change in changes.items():
                if isinstance(change, dict) and ("from" in change or "to" in change):
                    lines.append(f"{safe(field)}: {safe(change.get('from')) or '-'} -> {safe(change.get('to')) or '-'}")
                else:
                    lines.append(f"{safe(field)}: {safe(change)}")
            return "<br/>".join(lines)
    except (TypeError, json.JSONDecodeError):
        pass
    return safe(raw)


def generated_at():
    return datetime.now(timezone.utc).astimezone().strftime("%d/%m/%Y %H:%M:%S")


def excel_value(key, value):
    if value is None:
        return None
    if key not in {"opened_at", "deadline_at", "expected_close_at", "closed_at", "updated_at"}:
        return value
    raw = str(value).replace("Z", "+00:00")
    try:
        parsed = datetime.fromisoformat(raw)
        return parsed.replace(tzinfo=None)
    except ValueError:
        try:
            return datetime.strptime(str(value)[:10], "%Y-%m-%d")
        except ValueError:
            return value


def issues_xlsx(rows, applied_filters="Carteira completa acessível ao perfil"):
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Resumo"
    summary.sheet_view.showGridLines = False
    summary.merge_cells("A1:F2")
    summary["A1"] = "ATLAS - Relatório de Pendências FICO"
    summary["A1"].font = Font(name="Arial", size=18, bold=True, color="FFFFFF")
    summary["A1"].fill = PatternFill("solid", fgColor=BLUE)
    summary["A1"].alignment = Alignment(vertical="center")
    summary["A3"] = f"Gerado em {generated_at()}"
    summary["A3"].font = Font(name="Arial", size=9, color="697B8D")
    summary["A4"] = f"Recorte aplicado: {applied_filters}"
    summary.merge_cells("A4:F4")
    summary["A4"].font = Font(name="Arial", size=9, italic=True, color="697B8D")
    counts = Counter(text(row.get("status")) for row in rows)
    cards = [("Carteira total", len(rows)), ("Abertas", counts.get("Aberta", 0)), ("Em tratamento", counts.get("Em tratamento", 0)), ("Aguardando fiscal", counts.get("Aguardando validação", 0)), ("Baixadas", counts.get("Baixada", 0))]
    for column, (label, value) in enumerate(cards, start=1):
        cell = summary.cell(5, column, label)
        cell.font = Font(name="Arial", size=9, bold=True, color="697B8D")
        cell.fill = PatternFill("solid", fgColor=LIGHT)
        value_cell = summary.cell(6, column, value)
        value_cell.font = Font(name="Arial", size=18, bold=True, color=DARK)
        value_cell.fill = PatternFill("solid", fgColor=LIGHT)
        value_cell.alignment = Alignment(vertical="center")
    summary.row_dimensions[6].height = 32
    summary["A8"] = "Arquivo produzido pelo ATLAS. Os dados refletem a carteira acessível ao usuário no momento da exportação."
    summary.merge_cells("A8:F8")
    summary["A8"].font = Font(name="Arial", size=9, italic=True, color="697B8D")
    for column in range(1, 7):
        summary.column_dimensions[get_column_letter(column)].width = 23

    sheet = workbook.create_sheet("Pendências")
    sheet.sheet_view.showGridLines = False
    headers = ["ID ATLAS", "ID original", "Empresa", "Pacote", "Trecho", "Ativo", "Lado", "Especialidade", "Classificação", "Descrição", "KM inicial", "KM final", "Status", "Responsável FICO", "Responsável contratada", "Abertura", "Prazo", "Previsão de baixa", "Baixa", "Última atualização"]
    keys = ["id", "source_id", "company", "package", "segment", "asset", "side", "specialty", "classification", "description", "km_start", "km_end", "status", "fico_owner", "contractor_owner", "opened_at", "deadline_at", "expected_close_at", "closed_at", "updated_at"]
    sheet.append(headers)
    for row in rows:
        sheet.append([excel_value(key, row.get(key)) for key in keys])
    header_fill = PatternFill("solid", fgColor=BLUE)
    thin = Side(style="thin", color="DBE5EE")
    for cell in sheet[1]:
        cell.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial", size=9, color=DARK)
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=cell.column in (10,))
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for row in sheet.iter_rows(min_row=2, min_col=16, max_col=20):
        for cell in row:
            cell.number_format = "dd/mm/yyyy hh:mm" if cell.column == 20 else "dd/mm/yyyy"
    widths = [13, 13, 14, 14, 16, 22, 10, 22, 15, 55, 14, 14, 22, 22, 22, 14, 14, 18, 14, 21]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.row_dimensions[1].height = 32
    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def pdf_styles():
    styles = getSampleStyleSheet()
    return {
        "title": ParagraphStyle("AtlasTitle", parent=styles["Title"], fontName="Helvetica-Bold", fontSize=18, leading=22, textColor=colors.HexColor("#18324D"), alignment=TA_LEFT),
        "h2": ParagraphStyle("AtlasH2", parent=styles["Heading2"], fontName="Helvetica-Bold", fontSize=11, leading=14, textColor=colors.HexColor("#3578BC"), spaceBefore=6, spaceAfter=6),
        "body": ParagraphStyle("AtlasBody", parent=styles["BodyText"], fontName="Helvetica", fontSize=8.5, leading=11, textColor=colors.HexColor("#18324D")),
        "small": ParagraphStyle("AtlasSmall", parent=styles["BodyText"], fontName="Helvetica", fontSize=7, leading=9, textColor=colors.HexColor("#697B8D")),
        "center": ParagraphStyle("AtlasCenter", parent=styles["BodyText"], fontName="Helvetica", fontSize=8, leading=10, alignment=TA_CENTER),
    }


def header_story(title, subtitle, styles):
    logo = Image(str(LOGO), width=42 * mm, height=28 * mm) if LOGO.exists() else Paragraph("FICO", styles["title"])
    heading = [Paragraph(title, styles["title"]), Paragraph(subtitle, styles["small"])]
    table = Table([[logo, heading]], colWidths=[48 * mm, None])
    table.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "MIDDLE"), ("LINEBELOW", (0, 0), (-1, -1), 2, colors.HexColor("#F89B32")), ("BOTTOMPADDING", (0, 0), (-1, -1), 8)]))
    return [table, Spacer(1, 6 * mm)]


def issues_pdf(rows, applied_filters="Carteira completa acessível ao perfil"):
    stream = io.BytesIO()
    styles = pdf_styles()
    document = SimpleDocTemplate(stream, pagesize=landscape(A4), rightMargin=12 * mm, leftMargin=12 * mm, topMargin=10 * mm, bottomMargin=12 * mm, title="ATLAS - Relatório de Pendências")
    story = header_story("ATLAS - Relatório de Pendências", f"Carteira exportada em {generated_at()} - {len(rows)} registro(s)", styles)
    filter_box = Table([[Paragraph(f"<b>Recorte aplicado:</b> {safe(applied_filters)}", styles["small"])]], colWidths=[260 * mm])
    filter_box.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F5F8FB")), ("BOX", (0, 0), (-1, -1), .4, colors.HexColor("#DBE5EE")), ("LEFTPADDING", (0, 0), (-1, -1), 8), ("TOPPADDING", (0, 0), (-1, -1), 5), ("BOTTOMPADDING", (0, 0), (-1, -1), 5)]))
    story.extend([filter_box, Spacer(1, 4 * mm)])
    counts = Counter(text(row.get("status")) for row in rows)
    summary = [["Carteira", "Abertas", "Em tratamento", "Aguardando fiscal", "Baixadas"], [len(rows), counts.get("Aberta", 0), counts.get("Em tratamento", 0), counts.get("Aguardando validação", 0), counts.get("Baixada", 0)]]
    summary_table = Table(summary, colWidths=[36 * mm] * 5)
    summary_table.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E9F2F9")), ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#18324D")), ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"), ("FONTNAME", (0, 1), (-1, 1), "Helvetica-Bold"), ("FONTSIZE", (0, 0), (-1, -1), 8), ("ALIGN", (0, 0), (-1, -1), "CENTER"), ("BOX", (0, 0), (-1, -1), .5, colors.HexColor("#DBE5EE")), ("INNERGRID", (0, 0), (-1, -1), .25, colors.HexColor("#DBE5EE")), ("TOPPADDING", (0, 0), (-1, -1), 6), ("BOTTOMPADDING", (0, 0), (-1, -1), 6)]))
    story.extend([summary_table, Spacer(1, 6 * mm)])
    data = [["ID", "Empresa", "Especialidade", "Ativo / local", "Descrição", "Status", "Abertura", "Prazo", "Baixa"]]
    for row in rows:
        location = " / ".join(filter(None, [text(row.get("asset")), text(row.get("km_start")), text(row.get("km_end"))]))
        data.append([text(row.get("id")), text(row.get("company")), Paragraph(safe(row.get("specialty")), styles["small"]), Paragraph(safe(location), styles["small"]), Paragraph(safe(row.get("description")), styles["small"]), Paragraph(safe(row.get("status")), styles["small"]), date_text(row.get("opened_at")), date_text(row.get("deadline_at")), date_text(row.get("closed_at"))])
    table = LongTable(data, repeatRows=1, colWidths=[12 * mm, 20 * mm, 28 * mm, 35 * mm, 75 * mm, 30 * mm, 21 * mm, 21 * mm, 21 * mm])
    table.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#3578BC")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white), ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"), ("FONTSIZE", (0, 0), (-1, -1), 7), ("VALIGN", (0, 0), (-1, -1), "TOP"), ("GRID", (0, 0), (-1, -1), .25, colors.HexColor("#DBE5EE")), ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]), ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4)]))
    story.append(table)
    document.build(story, onFirstPage=pdf_footer, onLaterPages=pdf_footer)
    return stream.getvalue()


def pdf_footer(canvas, document):
    canvas.saveState()
    canvas.setFont("Helvetica", 7)
    canvas.setFillColor(colors.HexColor("#697B8D"))
    canvas.drawString(document.leftMargin, 7 * mm, "ATLAS - Desenvolvido por Thyago Viégas")
    canvas.drawRightString(document.pagesize[0] - document.rightMargin, 7 * mm, f"Página {document.page}")
    canvas.restoreState()


def bar_drawing(items, width=350, row_height=18, limit=8):
    selected = list(items)[:limit]
    height = max(1, len(selected)) * row_height
    drawing = Drawing(width, height)
    maximum = max((value for _, value in selected), default=1) or 1
    label_width = 115
    chart_width = width - label_width - 42
    for index, (label, value) in enumerate(selected):
        y = height - (index + 1) * row_height + 4
        drawing.add(String(0, y + 2, str(label)[:24], fontName="Helvetica", fontSize=7.5, fillColor=colors.HexColor("#18324D")))
        drawing.add(Rect(label_width, y, chart_width, 8, fillColor=colors.HexColor("#EDF2F6"), strokeColor=None, rx=4, ry=4))
        drawing.add(Rect(label_width, y, max(2, chart_width * value / maximum), 8, fillColor=colors.HexColor("#3578BC"), strokeColor=None, rx=4, ry=4))
        drawing.add(String(width - 34, y + 1, str(value), fontName="Helvetica-Bold", fontSize=8, fillColor=colors.HexColor("#18324D")))
    return drawing


def dashboard_pdf(rows, title="Dashboard executivo geral", applied_filters="Sem filtros adicionais"):
    stream = io.BytesIO()
    styles = pdf_styles()
    document = SimpleDocTemplate(stream, pagesize=landscape(A4), rightMargin=12 * mm, leftMargin=12 * mm, topMargin=9 * mm, bottomMargin=12 * mm, title=f"ATLAS - {title}")
    today = datetime.now().date().isoformat()
    total = len(rows)
    closed = sum(1 for row in rows if row.get("status") == "Baixada")
    backlog = sum(1 for row in rows if row.get("status") not in ("Baixada", "Cancelada"))
    waiting = sum(1 for row in rows if row.get("status") == "Aguardando validação")
    overdue = sum(1 for row in rows if row.get("status") not in ("Baixada", "Cancelada") and row.get("deadline_at") and str(row.get("deadline_at"))[:10] < today)
    closure_rate = round(closed / total * 100, 1) if total else 0
    story = header_story(title, f"Posição em {generated_at()} - Fonte: banco de dados ATLAS", styles)
    filter_box = Table([[Paragraph(f"<b>Recorte aplicado:</b> {safe(applied_filters)}", styles["small"])]], colWidths=[260 * mm])
    filter_box.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F5F8FB")), ("BOX", (0, 0), (-1, -1), .4, colors.HexColor("#DBE5EE")), ("LEFTPADDING", (0, 0), (-1, -1), 8), ("RIGHTPADDING", (0, 0), (-1, -1), 8), ("TOPPADDING", (0, 0), (-1, -1), 5), ("BOTTOMPADDING", (0, 0), (-1, -1), 5)]))
    story.extend([filter_box, Spacer(1, 4 * mm)])
    metrics = [("CARTEIRA TOTAL", total, "pendências no recorte"), ("BACKLOG ATIVO", backlog, "ainda não baixadas"), ("EM ATRASO", overdue, "prazo ultrapassado"), ("AGUARDANDO FISCAL", waiting, "prontas para validar"), ("TAXA DE BAIXA", f"{closure_rate}%", f"{closed} concluída(s)")]
    metric_cells = []
    for label, value, note in metrics:
        metric_cells.append([Paragraph(label, styles["small"]), Paragraph(f"<b><font size='18'>{value}</font></b>", styles["body"]), Paragraph(note, styles["small"])])
    metric_table = Table([metric_cells], colWidths=[52 * mm] * 5)
    metric_table.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, -1), colors.white), ("BOX", (0, 0), (-1, -1), .5, colors.HexColor("#DBE5EE")), ("INNERGRID", (0, 0), (-1, -1), .4, colors.HexColor("#DBE5EE")), ("VALIGN", (0, 0), (-1, -1), "MIDDLE"), ("TOPPADDING", (0, 0), (-1, -1), 8), ("BOTTOMPADDING", (0, 0), (-1, -1), 8)]))
    story.extend([metric_table, Spacer(1, 5 * mm)])

    status_counts = Counter(text(row.get("status")) or "Sem status" for row in rows).most_common()
    specialty_counts = Counter(text(row.get("specialty")) or "Sem especialidade" for row in rows).most_common()
    chart_left = [Paragraph("Distribuição por status", styles["h2"]), bar_drawing(status_counts, width=350, limit=7)]
    chart_right = [Paragraph("Pendências por especialidade", styles["h2"]), bar_drawing(specialty_counts, width=350, limit=7)]
    charts = Table([[chart_left, chart_right]], colWidths=[130 * mm, 130 * mm])
    charts.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP"), ("BOX", (0, 0), (-1, -1), .4, colors.HexColor("#DBE5EE")), ("INNERGRID", (0, 0), (-1, -1), .4, colors.HexColor("#DBE5EE")), ("LEFTPADDING", (0, 0), (-1, -1), 8), ("RIGHTPADDING", (0, 0), (-1, -1), 8), ("TOPPADDING", (0, 0), (-1, -1), 5), ("BOTTOMPADDING", (0, 0), (-1, -1), 5)]))
    story.extend([charts, Spacer(1, 4 * mm)])

    company_data = [["Empresa", "Carteira", "Baixadas", "Backlog", "Em atraso", "Taxa de baixa"]]
    by_company = {}
    for row in rows:
        by_company.setdefault(text(row.get("company")) or "Sem empresa", []).append(row)
    for company, company_rows in sorted(by_company.items(), key=lambda item: len(item[1]), reverse=True):
        company_total = len(company_rows)
        company_closed = sum(1 for row in company_rows if row.get("status") == "Baixada")
        company_backlog = sum(1 for row in company_rows if row.get("status") not in ("Baixada", "Cancelada"))
        company_overdue = sum(1 for row in company_rows if row.get("status") not in ("Baixada", "Cancelada") and row.get("deadline_at") and str(row.get("deadline_at"))[:10] < today)
        company_data.append([company, company_total, company_closed, company_backlog, company_overdue, f"{round(company_closed/company_total*100, 1) if company_total else 0}%"])
    company_table = Table(company_data, repeatRows=1, colWidths=[55 * mm, 34 * mm, 34 * mm, 34 * mm, 34 * mm, 34 * mm])
    company_table.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#3578BC")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white), ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"), ("FONTNAME", (0, 1), (-1, -1), "Helvetica"), ("FONTSIZE", (0, 0), (-1, -1), 8), ("ALIGN", (1, 0), (-1, -1), "CENTER"), ("GRID", (0, 0), (-1, -1), .3, colors.HexColor("#DBE5EE")), ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]), ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4)]))
    story.extend([Paragraph("Desempenho por empresa", styles["h2"]), company_table])
    document.build(story, onFirstPage=pdf_footer, onLaterPages=pdf_footer)
    return stream.getvalue()


def closure_certificate(issue, history, evidence):
    stream = io.BytesIO()
    styles = pdf_styles()
    document = SimpleDocTemplate(stream, pagesize=A4, rightMargin=16 * mm, leftMargin=16 * mm, topMargin=12 * mm, bottomMargin=15 * mm, title=f"Comprovante de encerramento ATLAS #{issue['id']}")
    canonical = json.dumps({"issue": dict(issue), "history": [dict(item) for item in history], "evidence": [dict(item) for item in evidence]}, ensure_ascii=False, default=str, sort_keys=True)
    verification = hashlib.sha256(canonical.encode("utf-8")).hexdigest().upper()
    story = header_story(f"Comprovante de encerramento #{issue['id']}", f"Documento de rastreabilidade emitido em {generated_at()}", styles)
    status_color = colors.HexColor("#E5F0F8")
    status = Table([["STATUS", issue.get("status"), "DATA DA BAIXA", date_text(issue.get("closed_at"))]], colWidths=[28 * mm, 45 * mm, 34 * mm, 45 * mm])
    status.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, -1), status_color), ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"), ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#18324D")), ("BOX", (0, 0), (-1, -1), .6, colors.HexColor("#3578BC")), ("VALIGN", (0, 0), (-1, -1), "MIDDLE"), ("TOPPADDING", (0, 0), (-1, -1), 8), ("BOTTOMPADDING", (0, 0), (-1, -1), 8)]))
    story.extend([status, Spacer(1, 5 * mm), Paragraph("Identificação da pendência", styles["h2"])])
    fields = [
        ("ID original", issue.get("source_id")), ("Empresa", issue.get("company")),
        ("Especialidade", issue.get("specialty")), ("Classificação", issue.get("classification")),
        ("Ativo", issue.get("asset")), ("Trecho", issue.get("segment")),
        ("KM inicial", issue.get("km_start")), ("KM final", issue.get("km_end")),
        ("Responsável FICO", issue.get("fico_owner")), ("Responsável contratada", issue.get("contractor_owner")),
        ("Abertura", date_text(issue.get("opened_at"))), ("Prazo", date_text(issue.get("deadline_at"))),
    ]
    rows = []
    for index in range(0, len(fields), 2):
        left, right = fields[index], fields[index + 1]
        rows.append([Paragraph(f"<b>{safe(left[0])}</b><br/>{safe(left[1]) or '-'}", styles["body"]), Paragraph(f"<b>{safe(right[0])}</b><br/>{safe(right[1]) or '-'}", styles["body"])])
    details = Table(rows, colWidths=[83 * mm, 83 * mm])
    details.setStyle(TableStyle([("GRID", (0, 0), (-1, -1), .35, colors.HexColor("#DBE5EE")), ("VALIGN", (0, 0), (-1, -1), "TOP"), ("TOPPADDING", (0, 0), (-1, -1), 6), ("BOTTOMPADDING", (0, 0), (-1, -1), 6)]))
    story.extend([details, Spacer(1, 4 * mm), Paragraph("Descrição", styles["h2"]), Paragraph(safe(issue.get("description")), styles["body"]), Spacer(1, 4 * mm), Paragraph("Histórico de movimentações e modificações", styles["h2"])])
    timeline = [["Data e hora", "Evento", "Transição", "Responsável", "Comentário / alterações"]]
    for item in history:
        transition = " -> ".join(filter(None, [text(item.get("from_status")), text(item.get("to_status"))]))
        timeline.append([Paragraph(safe(item.get("created_at")), styles["small"]), Paragraph(safe(item.get("event")), styles["small"]), Paragraph(safe(transition), styles["small"]), Paragraph(safe(item.get("actor")) or "Sistema", styles["small"]), Paragraph(history_comment(item.get("comment")), styles["small"])])
    timeline_table = LongTable(timeline, repeatRows=1, colWidths=[29 * mm, 32 * mm, 28 * mm, 33 * mm, 44 * mm])
    timeline_table.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#3578BC")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white), ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"), ("FONTSIZE", (0, 0), (-1, -1), 7), ("VALIGN", (0, 0), (-1, -1), "TOP"), ("GRID", (0, 0), (-1, -1), .3, colors.HexColor("#DBE5EE")), ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]), ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4)]))
    story.extend([timeline_table, Spacer(1, 4 * mm), Paragraph("Evidências vinculadas", styles["h2"])])
    evidence_data = [["Tipo", "Arquivo", "Captura", "Localização"]]
    for item in evidence:
        location = ", ".join(filter(None, [text(item.get("latitude")), text(item.get("longitude"))])) or "-"
        evidence_data.append([text(item.get("kind")), text(item.get("original_name")) or text(item.get("file_path")), text(item.get("captured_at")) or text(item.get("created_at")), location])
    evidence_table = Table(evidence_data, colWidths=[28 * mm, 70 * mm, 40 * mm, 28 * mm])
    evidence_table.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E9F2F9")), ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"), ("FONTSIZE", (0, 0), (-1, -1), 7), ("GRID", (0, 0), (-1, -1), .3, colors.HexColor("#DBE5EE")), ("VALIGN", (0, 0), (-1, -1), "TOP"), ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4)]))
    story.extend([evidence_table, Spacer(1, 6 * mm), Paragraph("Código de verificação do conteúdo", styles["h2"]), Paragraph(verification, ParagraphStyle("Hash", parent=styles["small"], fontName="Courier", wordWrap="CJK")), Spacer(1, 3 * mm), Paragraph("Este comprovante consolida os dados e a trilha de movimentações armazenados no ATLAS no momento da emissão. Alterações posteriores produzirão um código de verificação diferente.", styles["small"])])
    document.build(story, onFirstPage=pdf_footer, onLaterPages=pdf_footer)
    return stream.getvalue()
