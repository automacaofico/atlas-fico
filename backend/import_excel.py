import argparse
import json
import sys
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))
from server import db, initialize, iso, utcnow

DEFAULT_SOURCE = Path(r"C:\Users\engtv\Downloads\Planilhas\FICO-Vale\Banco de Dados - FICO - Entrega de Obras.xlsx")
SHEET = "Banco de Dados"


def clean(value):
    if value is None:
        return None
    text = str(value).strip()
    return text if text and text not in ("-", "nan", "NaT") else None


def date_value(value):
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = clean(value)
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def specialty(row):
    combined = " ".join(filter(None, [clean(row.get("Tipo de Protocolo")), clean(row.get("Tipo de Pendencia")), clean(row.get("Elemento"))])).lower()
    if "document" in combined or "databook" in combined or "rnc" in combined:
        return "Documental"
    if "dren" in combined or "bueiro" in combined or "assorea" in combined:
        return "Drenagem"
    if "terrap" in combined or "talude" in combined or "eros" in combined or "conforma" in combined:
        return "Terraplenagem"
    if any(x in combined for x in ("concreto", "estaca", "pilar", "laje", "obra de arte especial", "viga")):
        return "Estruturas"
    if "paviment" in combined or "sublastro" in combined:
        return "Pavimentação"
    return "Obras Complementares"


def status_value(value):
    return "Baixada" if clean(value) and clean(value).lower().startswith("encerr") else "Aberta"


def classification(value):
    value = clean(value)
    return value if value in ("Tipo A", "Tipo B", "Tipo C") else None


def read_rows(path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[SHEET]
    headers = [clean(x.value) for x in worksheet[2]]
    for excel_row, cells in enumerate(worksheet.iter_rows(min_row=3), start=3):
        row = dict(zip(headers, [cell.value for cell in cells]))
        substantive = any(clean(row.get(name)) for name in ("Empresa", "Ativo", "Descrição Pendencia", "Data_Abertura"))
        if substantive:
            yield excel_row, row


def import_workbook(path, apply=False):
    initialize()
    report = {"source": str(path), "sheet": SHEET, "mode": "apply" if apply else "dry-run", "read": 0, "valid": 0, "inserted": 0, "existing": 0, "rejected": 0, "errors": [], "specialties": {}, "companies": {}}
    with db() as connection:
        admin = connection.execute("SELECT id FROM users WHERE email='thyago.viegas@vale.com'").fetchone()[0]
        for excel_row, row in read_rows(path):
            report["read"] += 1
            source_id = row.get("Id_Pendencia")
            company_name = clean(row.get("Empresa"))
            description = clean(row.get("Descrição Pendencia"))
            asset = clean(row.get("Ativo"))
            cls = classification(row.get("Classificação da Pendencia"))
            opened = date_value(row.get("Data_Abertura"))
            problems = []
            if source_id is None: problems.append("Id_Pendencia ausente")
            if not company_name: problems.append("Empresa ausente")
            if not description: problems.append("Descrição ausente")
            if not asset: problems.append("Ativo ausente")
            if not cls: problems.append("Classificação inválida")
            if not opened: problems.append("Data_Abertura inválida")
            if problems:
                report["rejected"] += 1
                report["errors"].append({"excel_row": excel_row, "source_id": source_id, "problems": problems})
                continue
            report["valid"] += 1
            spec = specialty(row)
            report["specialties"][spec] = report["specialties"].get(spec, 0) + 1
            report["companies"][company_name] = report["companies"].get(company_name, 0) + 1
            if connection.execute("SELECT 1 FROM issues WHERE source_id=?", (source_id,)).fetchone():
                report["existing"] += 1
                continue
            if not apply:
                continue
            company = connection.execute("SELECT id FROM companies WHERE name=?", (company_name,)).fetchone()
            if not company:
                cursor = connection.execute("INSERT INTO companies(name,kind) VALUES(?,?)", (company_name, "CONTRATADA"))
                company_id = cursor.lastrowid
            else:
                company_id = company[0]
            cursor = connection.execute(
                "INSERT INTO issues(source_id,package,segment,asset,side,company_id,protocol_code,origin,protocol,protocol_type,protocol_item,element,specialty,description,classification,km_start,km_end,status,contractor_owner,fico_owner,opened_at,deadline_at,expected_close_at,closed_at,notes,created_by,updated_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                (
                    source_id, clean(row.get("Pacote")), clean(row.get("Segmento")), asset, clean(row.get("Lado")), company_id,
                    clean(row.get("Cod_Protocolo")), clean(row.get("Origem")), clean(row.get("Protocolo")), clean(row.get("Tipo de Protocolo")),
                    clean(row.get("Item do Protocolo")), clean(row.get("Elemento")), spec, description, cls, clean(row.get("KM Inicial")), clean(row.get("KM Final")),
                    status_value(row.get("Status_Pendencia")), clean(row.get("Responsável Contratada")), clean(row.get("Responsável Vale")), opened,
                    date_value(row.get("Data_Prazo")), date_value(row.get("Data_Prevista Encerramento")), date_value(row.get("Data_Encerramento")),
                    clean(row.get("Obs:")), admin, iso(utcnow()),
                ),
            )
            issue_id = cursor.lastrowid
            connection.execute("INSERT INTO issue_history(issue_id,event,to_status,comment,actor_id) VALUES(?,?,?,?,?)", (issue_id, "IMPORTADO_EXCEL", status_value(row.get("Status_Pendencia")), f"Importado da linha {excel_row} da base unificada", admin))
            report["inserted"] += 1
    report_path = Path(__file__).resolve().parent / "data" / "import-report.json"
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({k: report[k] for k in ("mode", "read", "valid", "inserted", "existing", "rejected")}, ensure_ascii=False, indent=2))
    print(f"Relatório: {report_path}")
    return 0 if report["rejected"] == 0 else 2


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Valida e importa a base unificada FICO no ATLAS")
    parser.add_argument("source", nargs="?", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument("--apply", action="store_true", help="Grava registros válidos no banco de homologação")
    args = parser.parse_args()
    raise SystemExit(import_workbook(args.source, args.apply))
